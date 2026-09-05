#!/usr/bin/env python
"""同步 MaaFGO 礼装 list/team 高价值识别资源。"""

from __future__ import annotations

import argparse
import hashlib
import shutil
import tempfile
from pathlib import Path
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from PIL import Image


DEFAULT_ICON_ROOT = Path(r"F:\Game_AUTO\MaaFGO\FGO整包\图标_游戏内显示")
DEFAULT_CONFIG = Path(r"F:\KumikoPythonCode\FGO主线剧情epub\equip_list.xlsx")
PROJECT_ROOT = Path(__file__).resolve().parents[3]


def selected_filenames(config: Path) -> set[str]:
    workbook = load_workbook(config, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    indices = {name: index for index, name in enumerate(headers)}
    missing_columns = {"equipId", "非高价值礼装"} - indices.keys()
    if missing_columns:
        raise ValueError(f"配置表缺少列：{', '.join(sorted(missing_columns))}")

    selected = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        flag = row[indices["非高价值礼装"]]
        if flag == 0 or str(flag).strip() == "0":
            selected.add(f"f_{int(row[indices['equipId']])}0.png")
    return selected


def png_names(folder: Path) -> set[str]:
    return {path.name for path in folder.glob("*.png")}


def save_png(image: Image.Image, target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(dir=target.parent, suffix=".tmp", delete=False) as file:
        temporary = Path(file.name)
    try:
        image.save(temporary, format="PNG", optimize=True, compress_level=9)
        temporary.replace(target)
    finally:
        temporary.unlink(missing_ok=True)


def download_face(filename: str, faces: Path) -> None:
    target = faces / filename
    if target.exists():
        return
    request = Request(
        f"https://static.atlasacademy.io/JP/Faces/{filename}",
        headers={"User-Agent": "MaaFGO equip icon pipeline"},
    )
    with urlopen(request, timeout=30) as response:
        payload = response.read()
    if not payload.startswith(b"\x89PNG\r\n\x1a\n"):
        raise ValueError(f"Atlas 返回的不是 PNG：{filename}")
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(payload)
    try:
        with Image.open(target) as image:
            if image.size != (128, 128) or image.mode != "RGBA":
                raise ValueError(f"Atlas Faces 格式异常：{filename} {image.size} {image.mode}")
    except Exception:
        target.unlink(missing_ok=True)
        raise


def build_list_icon(source: Path, target: Path) -> None:
    with Image.open(source) as image:
        if image.size != (128, 128) or image.mode != "RGBA":
            raise ValueError(f"Faces 格式异常：{source.name} {image.size} {image.mode}")
        resized = image.resize((158, 158), Image.Resampling.BILINEAR)
        quantized = resized.quantize(
            colors=256,
            method=Image.Quantize.FASTOCTREE,
            dither=Image.Dither.NONE,
        )
        save_png(quantized.crop((11, 27, 158, 83)), target)


def build_team_icon(source: Path, target: Path) -> None:
    with Image.open(source) as image:
        if image.width != 188 or image.height < 86:
            raise ValueError(f"188 源图格式异常：{source.name} {image.size}")
        top = (image.height - 86) // 2
        cropped = image.convert("RGBA").crop((18, top, 170, top + 86))
        quantized = cropped.quantize(
            colors=256,
            method=Image.Quantize.FASTOCTREE,
            dither=Image.Dither.NONE,
        )
        save_png(quantized, target)


def sync_selected(source: Path, target: Path, selected: set[str], dry_run: bool) -> None:
    available = png_names(source)
    missing = selected - available
    if missing:
        raise FileNotFoundError(f"{source} 缺少 {len(missing)} 张：{', '.join(sorted(missing)[:5])}")
    existing = png_names(target) if target.exists() else set()
    if dry_run:
        print(f"{target}: copy={len(selected - existing)}, remove={len(existing - selected)}")
        return
    target.mkdir(parents=True, exist_ok=True)
    for filename in selected:
        shutil.copy2(source / filename, target / filename)
    for stale in target.glob("*.png"):
        if stale.name not in selected:
            stale.unlink()


def sha256(path: Path) -> str:
    with path.open("rb") as file:
        return hashlib.file_digest(file, "sha256").hexdigest()


def verify(source: Path, target: Path, expected: set[str], size: tuple[int, int]) -> None:
    if png_names(source) & expected != expected or png_names(target) != expected:
        raise AssertionError(f"文件集合校验失败：{target}")
    for filename in expected:
        if sha256(source / filename) != sha256(target / filename):
            raise AssertionError(f"哈希校验失败：{filename}")
        with Image.open(target / filename) as image:
            if image.size != size or image.mode != "P":
                raise AssertionError(f"图片格式校验失败：{filename} {image.size} {image.mode}")


def sync_list(icon_root: Path, project_root: Path, selected: set[str], args: argparse.Namespace) -> None:
    faces = icon_root / "Faces"
    full = icon_root / "Faces_礼装_158_256色无抖动_上裁27左裁11下裁75_仓库界面识别用"
    high = icon_root / f"{full.name}_高价值礼装"
    needed = selected - png_names(full)
    if needed and args.high_value_only:
        raise FileNotFoundError(f"list 完整识别资源缺少 {len(needed)} 张；不要使用 --high-value-only")
    for filename in sorted(needed):
        raw = faces / filename
        if not raw.exists():
            if not args.download_missing:
                raise FileNotFoundError(f"Faces 缺少 {filename}；加 --download-missing 后可从 Atlas 补图")
            if not args.dry_run:
                download_face(filename, faces)
        if not args.dry_run:
            build_list_icon(raw, full / filename)
    sync_selected(full, high, selected, args.dry_run)
    deploy = project_root / "assets" / "resource" / "base" / "image" / "EquipFaces" / "list"
    sync_selected(high, deploy, selected, args.dry_run)
    if not args.dry_run:
        verify(high, deploy, selected, (147, 56))


def sync_team(icon_root: Path, project_root: Path, selected: set[str], args: argparse.Namespace) -> None:
    source = icon_root / "EquipFaces_188"
    full = icon_root / "EquipFaces_188_左右裁18_上下居中保留86_256色无抖动_编队用"
    high = icon_root / f"{full.name}_高价值礼装"
    source_names = png_names(source)
    missing = selected - source_names
    if missing:
        raise FileNotFoundError(f"EquipFaces_188 缺少 {len(missing)} 张：{', '.join(sorted(missing)[:5])}")
    if not args.high_value_only:
        if args.dry_run:
            print(f"{full}: regenerate={len(source_names)}")
        else:
            for filename in sorted(source_names):
                build_team_icon(source / filename, full / filename)
            for stale in full.glob("*.png"):
                if stale.name not in source_names:
                    stale.unlink()
    sync_selected(full, high, selected, args.dry_run)
    deploy = project_root / "assets" / "resource" / "base" / "image" / "EquipFaces" / "team"
    sync_selected(high, deploy, selected, args.dry_run)
    if not args.dry_run:
        verify(high, deploy, selected, (152, 86))


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--target", choices=("list", "team", "all"), default="all")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--download-missing", action="store_true")
    parser.add_argument("--high-value-only", action="store_true")
    parser.add_argument("--icon-root", type=Path, default=DEFAULT_ICON_ROOT)
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG)
    parser.add_argument("--project-root", type=Path, default=PROJECT_ROOT)
    args = parser.parse_args()

    selected = selected_filenames(args.config)
    print(f"高价值礼装：{len(selected)}")
    if args.target in ("list", "all"):
        sync_list(args.icon_root, args.project_root, selected, args)
    if args.target in ("team", "all"):
        sync_team(args.icon_root, args.project_root, selected, args)


if __name__ == "__main__":
    main()
